import sqlite3
import pandas as pd
import datetime
from openpyxl import Workbook
import smtplib
from email.message import EmailMessage
import datetime
with open("/tmp/send_attendance_log.txt", "a") as f:
    f.write(f"Script started at {datetime.datetime.now()}\n")

# 参数配置（请按你实际情况修改）
DB_PATH = '/home/ec2-user/Student-Checkin-System/student-checkin-backend/data/student_checkin_system_imported.db'
REPORT_FOLDER = '/home/ec2-user/Student-Checkin-System/reports/'

# 邮件配置
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 587
SENDER_EMAIL = 'studentattendancereport49@gmail.com'  # 发送人邮箱
SENDER_PASSWORD = 'mxblqmkeeelbvuyn'  # Google生成的"应用专用密码"
RECEIVER_EMAILS = ['studentattendancereport49@gmail.com']  # 收件人邮箱

# 获取今天日期
today = datetime.date.today().isoformat()

# 生成报表
conn = sqlite3.connect(DB_PATH)

query = f"""
SELECT s.grade, s.name AS student_name, s.father_name, s.mother_name, s.phone_number,
       ci.time AS checkin_time
FROM students s
LEFT JOIN (SELECT * FROM checkins WHERE DATE(time) = DATE('{today}')) ci ON s.id = ci.student_id
ORDER BY s.grade, s.name
"""

df = pd.read_sql_query(query, conn)
conn.close()

def attendance_status(row):
    return "出勤" if row['checkin_time'] else "缺勤"

df['出勤状态'] = df.apply(attendance_status, axis=1)

# 生成Excel报表
report_filename = f'Attendance_{today}.xlsx'
output_path = REPORT_FOLDER + report_filename

wb = Workbook()
for grade, group in df.groupby('grade'):
    ws = wb.create_sheet(title=grade)
    ws.append(['学生姓名', '家长1', '家长2', '电话', '出勤状态'])
    for _, row in group.iterrows():
        ws.append([
            row['student_name'], row['father_name'], row['mother_name'], row['phone_number'], row['出勤状态']
        ])
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
wb.save(output_path)

print(f'✅ 出勤报表已生成: {output_path}')

# 发送邮件
msg = EmailMessage()
msg['Subject'] = f"学生出勤报表 - {today}"
msg['From'] = SENDER_EMAIL
msg['To'] = ', '.join(RECEIVER_EMAILS)
msg.set_content(f"老师好，\n\n附件为 {today} 学生出勤报表，请查收。\n\n祝好！\n")

# 添加附件
with open(output_path, 'rb') as f:
    file_data = f.read()
    msg.add_attachment(file_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=report_filename)

# 发送邮件
with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
    server.starttls()
    server.login(SENDER_EMAIL, SENDER_PASSWORD)
    server.send_message(msg)

print("✅ 邮件已成功发送给老师")
